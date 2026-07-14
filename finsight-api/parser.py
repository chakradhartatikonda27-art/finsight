"""
FinSight Tally Parser — Mudduluru Infratech KA
Correctly parses the 4-row-per-voucher Tally Day Book format.

Column structure (Row 8 = header):
Col 0: Date
Col 1: Particulars (Ledger Name)
Col 2: Amount (Dr side sub-entry) or ''
Col 3: 'Dr' or 'Cr' marker
Col 4: Vch Type
Col 5: Vch No.
Col 6: Debit Amount
Col 7: Credit Amount
"""
import openpyxl
import xlrd
import csv
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from typing import Optional
from io import BytesIO
import os

def to_decimal(val) -> Decimal:
    try:
        if val is None or str(val).strip() in ('', '—', '-', 'N/A'):
            return Decimal('0.00')
        clean = str(val).replace(',', '').strip()
        if clean.startswith('(') and clean.endswith(')'):
            clean = '-' + clean[1:-1]
        return Decimal(str(clean)).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return Decimal('0.00')

def parse_date_str(val) -> Optional[date]:
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ['%d-%m-%Y','%d/%m/%Y','%Y-%m-%d','%d-%b-%Y','%d %b %Y','%d-%b-%y']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def load_wb(filepath: str):
    with open(filepath, 'rb') as f:
        content = f.read()
    if content[:4] == b'PK\x03\x04':
        return openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    raise ValueError("Not a ZIP/XLSX file")

def parse_xls(filepath: str) -> dict:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
        return _extract_simple(rows)

    wb = load_wb(filepath)
    print(f"Sheets: {wb.sheetnames}")

    # Parse Ledger Codes sheet
    ledger_map = {}
    if 'Ledger Codes' in wb.sheetnames:
        ws = wb['Ledger Codes']
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or row[6] is None:
                continue
            ledger_name = str(row[6]).strip()
            ledger_code = str(row[11]).strip() if len(row) > 11 and row[11] else ''
            level0 = str(row[2]).strip() if row[2] else ''
            level1 = str(row[3]).strip() if row[3] else ''
            level2 = str(row[4]).strip() if row[4] else ''
            level3 = str(row[5]).strip() if row[5] else ''
            statement = str(row[1]).strip() if row[1] else ''
            if ledger_name:
                ledger_map[ledger_name] = {
                    'ledger_code': ledger_code,
                    'statement':   statement,
                    'tally_group': level2 or level1 or level0,
                    'level0': level0, 'level1': level1,
                    'level2': level2, 'level3': level3,
                }

    # Parse Opening Balances sheet
    opening_map = {}
    if 'Opening bal' in wb.sheetnames:
        ws = wb['Opening bal']
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            dr = to_decimal(row[1] if len(row) > 1 else None)
            cr = to_decimal(row[2] if len(row) > 2 else None)
            if name:
                opening_map[name] = {'opening_dr': float(dr), 'opening_cr': float(cr)}

    # Parse Day Book sheet
    if 'Day Book' not in wb.sheetnames:
        wb.close()
        raise ValueError("No 'Day Book' sheet found")

    ws = wb['Day Book']
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    return _extract_tally_vouchers(rows, ledger_map, opening_map)


def _extract_tally_vouchers(rows, ledger_map: dict, opening_map: dict) -> dict:
    """
    Parse Tally's 4-row-per-voucher Day Book format.
    
    Header row (Row 8, index 8):
    Date | Particulars | '' | '' | Vch Type | Vch No. | Debit Amount | Credit Amount
    
    Voucher entry rows:
    Row A: date | ledger_dr | '' | '' | vch_type | vch_no | debit_amt | None
    Row B: None | cost_centre | amt | 'Dr' | None | None | None | None
    Row C: ''   | ledger_cr | '' | '' | '' | '' | None | credit_amt
    Row D: None | cost_centre | amt | 'Cr' | None | None | None | None
    """
    total_rows = len(rows)
    vouchers = []
    ledgers = {}

    # Find header row
    header_row = 8
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == 'Date':
            header_row = i
            break

    current_date = None
    current_vch_type = None
    current_vch_no = None
    current_cost_centre = None

    i = header_row + 1
    while i < len(rows):
        row = rows[i]
        if not row:
            i += 1
            continue

        def c(idx):
            if idx >= len(row) or row[idx] is None:
                return ''
            return str(row[idx]).strip()

        col0 = row[0]
        col1 = c(1)
        col3 = c(3)
        col4 = c(4)
        col5 = c(5)
        col6 = row[6] if len(row) > 6 else None
        col7 = row[7] if len(row) > 7 else None

        # Skip empty rows
        if not any(row[j] for j in range(len(row)) if row[j] is not None):
            i += 1
            continue

        # Date detection
        if col0 is not None and col0 != '':
            d = parse_date_str(col0)
            if d:
                current_date = d

        # Voucher type and number
        if col4 and col4.lower() not in ('vch type', 'type', ''):
            current_vch_type = col4
        if col5 and col5.lower() not in ('vch no.', 'vch no', 'no', ''):
            current_vch_no = col5

        # Cost centre from Dr/Cr marker rows (col3 == 'Dr' or 'Cr')
        if col3 in ('Dr', 'Cr') and col1:
            current_cost_centre = col1

        # Skip Dr/Cr marker rows and header-like rows
        if col3 in ('Dr', 'Cr'):
            i += 1
            continue

        if col1.lower() in ('particulars', 'date', ''):
            i += 1
            continue

        # Extract amounts
        debit  = to_decimal(col6)
        credit = to_decimal(col7)

        if col1 and current_date and (debit > 0 or credit > 0):
            # Clean ledger name — remove code suffix like (KA-1234567)
            ledger_name = col1
            if '(' in ledger_name and ledger_name.endswith(')'):
                clean_name = ledger_name[:ledger_name.rfind('(')].strip()
                ledger_code_from_name = ledger_name[ledger_name.rfind('(')+1:-1]
            else:
                clean_name = ledger_name
                ledger_code_from_name = ''

            vouchers.append({
                'txn_date':     str(current_date),
                'voucher_type': current_vch_type or 'Journal',
                'voucher_no':   current_vch_no   or f'AUTO/{i}',
                'ledger_name':  clean_name,
                'debit':        float(debit),
                'credit':       float(credit),
                'cost_centre':  current_cost_centre,
                'narration':    None,
            })

            if clean_name not in ledgers:
                lm = ledger_map.get(clean_name, ledger_map.get(ledger_name, {}))
                ob = opening_map.get(clean_name, opening_map.get(ledger_name, {}))
                ledgers[clean_name] = {
                    'ledger_name':  clean_name,
                    'ledger_code':  lm.get('ledger_code', ledger_code_from_name),
                    'tally_group':  lm.get('tally_group', 'Unknown'),
                    'statement':    lm.get('statement', ''),
                    'level0':       lm.get('level0', ''),
                    'level1':       lm.get('level1', ''),
                    'level2':       lm.get('level2', ''),
                    'opening_dr':   ob.get('opening_dr', 0.0),
                    'opening_cr':   ob.get('opening_cr', 0.0),
                }

        i += 1

    return {
        'total_rows':    total_rows,
        'voucher_count': len(vouchers),
        'ledger_count':  len(ledgers),
        'vouchers':      vouchers,
        'ledgers':       list(ledgers.values()),
    }


def _extract_simple(rows) -> dict:
    """Fallback simple parser for CSV files."""
    return _extract_tally_vouchers(rows, {}, {})


def save_to_supabase(parsed: dict, org_id: str, upload_id: str, supabase_client) -> dict:
    sb = supabase_client
    vouchers = parsed['vouchers']
    inserted = 0
    for i in range(0, len(vouchers), 200):
        batch = [dict(v, org_id=org_id, upload_id=upload_id) for v in vouchers[i:i+200]]
        sb.table('vouchers').insert(batch).execute()
        inserted += len(batch)
    for ledger in parsed['ledgers']:
        sb.table('ledger_mappings').upsert({
            'org_id':       org_id,
            'ledger_name':  ledger['ledger_name'],
            'tally_group':  ledger.get('tally_group', 'Unknown'),
            'mis_head':     None,
            'confidence':   0,
            'confirmed':    False,
        }, on_conflict='org_id,ledger_name').execute()
    return {'vouchers_inserted': inserted, 'ledgers_upserted': len(parsed['ledgers'])}
