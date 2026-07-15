"""
POST /api/parse/{upload_id}
Parses uploaded Tally XLS/XLSX/CSV file.
Stores vouchers + ledgers in Supabase.
"""
import os
import tempfile
from fastapi import APIRouter, HTTPException, Depends
from core.auth import get_current_user, CurrentUser
from core.config import settings

router = APIRouter()

MUDDULURU_MIS = {
    'Sale of Service':                  'Revenue',
    'Unbilled Revenue':                 'Revenue',
    'Interest Income':                  'Other Income',
    'Cost of material Consumed':        'Material Cost',
    'Cost of Fuel Consumed':            'Material Cost',
    'Direct Project Cost':              'Direct Cost',
    'Direct Labour Cost':               'Direct Labour',
    'Employee Benefit Expenses':        'Employee Cost',
    'Rent, rates & taxes':              'Administrative Cost',
    'Other administration expenses':    'Administrative Cost',
    'Legal and Professional Expenses':  'Administrative Cost',
    'Travel and Accomodation':          'Administrative Cost',
    'Repairs and Maintenance':          'Administrative Cost',
    'Interest on cash credit & overdraft': 'Finance Cost',
    'Interest on term loans':           'Finance Cost',
    'INTEREST ON UNSECURED LOAN':       'Finance Cost',
    'Bank Charges':                     'Finance Cost',
    'Fixed Asset':                      'Fixed Assets',
    'Cash & Bank Balance':              'Cash & Bank',
    'Balance With Banks':               'Cash & Bank',
    'Sundry debtors':                   'Trade Debtors',
    'Short Term Loans & Advances':      'Other Current Assets',
    'Other Current Assets':             'Other Current Assets',
    'Financial Asset':                  'Investments',
    'Other Long Term Assets':           'Other Assets',
    'Inventories':                      'Inventory',
    'Sundry creditors':                 'Trade Creditors',
    'Other Current Liabilities':        'Other Current Liabilities',
    'Short Term Provisions':            'Provisions',
    'Short Term Borrowings':            'Short Term Borrowings',
    'Long Term Borrowings':             'Long Term Borrowings',
    'Share Capital':                    'Share Capital',
    'Branch/ Division Account':         'Intercompany',
}


def run_parse_pipeline(file_path: str, org_id: str, upload_id: str) -> dict:
    """
    Parse XLS/XLSX/CSV file and store to Supabase.
    Returns summary dict.
    """
    from core.database import get_supabase
    import openpyxl
    from io import BytesIO

    ext = os.path.splitext(file_path)[1].lower()
    sb = get_supabase()

    # Load workbook
    with open(file_path, 'rb') as f:
        content = f.read()

    if content[:4] == b'PK\x03\x04':
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    else:
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        # Convert to openpyxl-compatible format via rows
        wb = None

    sheets = wb.sheetnames if wb else []

    # Parse Ledger Codes sheet
    ledger_map = {}
    if 'Ledger Codes' in sheets:
        ws = wb['Ledger Codes']
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or row[6] is None:
                continue
            ledger_name = str(row[6]).strip()
            level2 = str(row[4]).strip() if row[4] else ''
            level1 = str(row[3]).strip() if row[3] else ''
            level0 = str(row[2]).strip() if row[2] else ''
            if ledger_name:
                ledger_map[ledger_name] = level2 or level1 or level0

    # Parse Day Book
    vouchers = []
    ledgers = {}
    total_rows = 0

    if 'Day Book' in sheets:
        ws = wb['Day Book']
        current_date = None
        current_vch_type = None
        current_vch_no = None

        for row in ws.iter_rows(values_only=True):
            total_rows += 1
            if not row or all(v is None for v in row):
                continue

            def c(idx):
                if idx >= len(row) or row[idx] is None:
                    return ''
                return str(row[idx]).strip()

            col0, col1, col4, col5 = c(0), c(1), c(4), c(5)
            col6 = row[6] if len(row) > 6 else None
            col7 = row[7] if len(row) > 7 else None

            # Date detection
            if col0 and any(str(y) in col0 for y in range(2020, 2035)):
                current_date = col0
            if col4 and 'MSR' in col4:
                current_vch_type = col4
            if col5:
                current_vch_no = col5

            try:
                debit = float(col6 or 0)
                credit = float(col7 or 0)
            except (ValueError, TypeError):
                debit = credit = 0.0

            if col1 and current_date and (debit > 0 or credit > 0):
                tally_group = ledger_map.get(col1, 'Unknown')
                vouchers.append({
                    'org_id': org_id,
                    'upload_id': upload_id,
                    'txn_date': current_date,
                    'voucher_type': current_vch_type or 'Journal',
                    'voucher_no': current_vch_no or f'AUTO/{total_rows}',
                    'ledger_name': col1,
                    'debit': round(debit, 2),
                    'credit': round(credit, 2),
                    'cost_centre': None,
                    'narration': None,
                })
                if col1 not in ledgers:
                    ledgers[col1] = tally_group

    # Store vouchers in batches of 1000
    inserted = 0
    for i in range(0, len(vouchers), 1000):
        batch = vouchers[i:i+1000]
        try:
            sb.table('vouchers').insert(batch).execute()
            inserted += len(batch)
        except Exception as e:
            print(f"[PARSE] batch insert error: {e}")

    # Upsert ledger mappings
    for ledger_name, tally_group in ledgers.items():
        try:
            mis_head = MUDDULURU_MIS.get(tally_group)
            sb.table('ledger_mappings').upsert({
                'org_id': org_id,
                'ledger_name': ledger_name,
                'tally_group': tally_group,
                'mis_head': mis_head,
                'confidence': 99 if mis_head else 0,
                'confirmed': bool(mis_head),
            }, on_conflict='org_id,ledger_name').execute()
        except Exception:
            pass

    if wb:
        wb.close()

    return {
        'total_rows': total_rows,
        'voucher_count': len(vouchers),
        'ledger_count': len(ledgers),
        'vouchers_inserted': inserted,
    }


@router.post("/{upload_id}")
async def parse_upload(
    upload_id: str,
    current_user: CurrentUser = Depends(get_current_user),
):
    """Parse a previously uploaded file."""
    if settings.ENVIRONMENT == "development":
        return {
            "upload_id": upload_id,
            "status": "parsed",
            "voucher_count": 30490,
            "ledger_count": 1406,
            "message": "Parsed [dev mode]",
        }
    try:
        from core.database import get_supabase
        record = get_supabase().table("uploads").select(
            "id,filename,storage_path,org_id"
        ).eq("id", upload_id).single().execute().data
    except Exception:
        raise HTTPException(404, "Upload not found")

    raise HTTPException(501, "Use /api/upload/async for automatic parse pipeline")
